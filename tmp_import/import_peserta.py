#!/usr/bin/env python3
"""
Import peserta PORJAR dari folder xlsx ke database.
Membuat: schools, teams, users (player), team_members
"""

import os
import re
import uuid
import secrets
import string
import psycopg2
from datetime import datetime
import openpyxl
import bcrypt

# ── DB config ──────────────────────────────────────────────────────────────────
DB = dict(host='172.24.0.2', port=5432, dbname='porjar',
          user='porjar', password='o12S1kJiScjmejzPcIoAgyg9MUq9SD6f')

BASE_DIR = '/home/oman/project/porjar/tmp_import/peserta porjar'

# ── Mapping folder → (game_slug, tingkat, nomor_pertandingan_display) ──────────
FOLDER_MAP = {
    'EFO DUO SD':               ('efootball-duo',  'SD',  'eFootball Duo'),
    'EFO DUO SMA':              ('efootball-duo',  'SMA', 'eFootball Duo'),
    'EFO DUO SMP':              ('efootball-duo',  'SMP', 'eFootball Duo'),
    'EFO SOLO SD':              ('efootball-solo', 'SD',  'eFootball Solo'),
    'EFO SOLO SMA':             ('efootball-solo', 'SMA', 'eFootball Solo'),
    'EFO SOLO SMP':             ('efootball-solo', 'SMP', 'eFootball Solo'),
    'FREE FIRE SD':             ('ff',             'SD',  'Free Fire'),
    'FREE FIRE SMA LAKI-LAKI':  ('ff',             'SMA', 'Free Fire'),
    'FREE FIRE SMA PEREMPUAN':  ('ff',             'SMA', 'Free Fire'),
    'FREE FIRE SMP LAKI-LAKI':  ('ff',             'SMP', 'Free Fire'),
    'FREE FIRE SMP PEREMPUAN':  ('ff',             'SMP', 'Free Fire'),
    'HOK SMA':                  ('hok',            'SMA', 'HOK'),
    'HOK SMP':                  ('hok',            'SMP', 'HOK'),
    'MOBILE LEGENDS SD':        ('ml-pria',        'SD',  'ML Pria'),
    'MOBILE LEGENDS SMA LADIES':('ml-wanita',      'SMA', 'ML Wanita'),
    'MOBILE LEGENDS SMA LAKI-LAKI': ('ml-pria',   'SMA', 'ML Pria'),
    'MOBILE LEGENDS SMP LADIES':('ml-wanita',      'SMP', 'ML Wanita'),
    'MOBILE LEGENDS SMP LAKI-LAKI': ('ml-pria',   'SMP', 'ML Pria'),
    'PUBGM SMA':                ('pubgm',           'SMA', 'PUBG Mobile'),
    'PUBGM SMP':                ('pubgm',           'SMP', 'PUBG Mobile'),
}

# ── Helpers ────────────────────────────────────────────────────────────────────

def clean_str(v):
    """Strip whitespace, remove leading quotes/apostrophes."""
    if v is None:
        return ''
    s = str(v).strip()
    # Remove leading single-quote (sometimes Google Forms exports do this)
    s = s.lstrip("'`'")
    return s.strip()

def clean_nik(v):
    """Normalize NIK to string of digits only."""
    s = clean_str(v)
    # Remove non-digit chars
    s = re.sub(r'\D', '', s)
    return s if s else None

def gen_password(length=8):
    alphabet = string.ascii_letters + string.digits
    return ''.join(secrets.choice(alphabet) for _ in range(length))

def hash_password(plain):
    return bcrypt.hashpw(plain.encode(), bcrypt.gensalt(rounds=10)).decode()

def col_idx(headers, *names):
    """Find first matching column index (case-insensitive)."""
    headers_lower = [h.lower().strip() if h else '' for h in headers]
    for name in names:
        try:
            return headers_lower.index(name.lower())
        except ValueError:
            pass
    return None

def extract_players(ws, folder_name):
    """
    Parse worksheet rows → list of teams.
    Returns: list of dicts {team_name, school_name, phone, players: [{name, nik, is_captain}]}
    """
    headers = [cell.value for cell in ws[1]]

    # Detect school column - SD uses DAERAH, others use NAMA SEKOLAH
    school_col = col_idx(headers, 'nama sekolah', 'daerah', 'nama sekolah/daerah')
    team_col   = col_idx(headers, 'nama tim')
    phone_col  = col_idx(headers, 'no telp/wa (pembina/kapten)', 'no telp/wa (pembina)', 'no telp/wa')

    # Find player columns: groups of (nama, nik) per player
    # Pattern: NAMA LENGKAP PEMAIN N + NIK N
    player_groups = []  # list of (nama_col, nik_col, is_captain)

    # Scan headers to find player name + NIK pairs
    i = 0
    player_num = 0
    while i < len(headers):
        h = (headers[i] or '').strip().upper()
        # Match player name columns
        if re.match(r'NAMA (LENGKAP )?(PEMAIN|PLAYER)', h) or \
           re.match(r'NAMA (PEMAIN|PLAYER)', h) or \
           (player_num == 0 and 'KAPTEN' in h and 'NAMA' in h):
            # Next, find NIK column for this player
            # NIK might be at i+1 or i+2 depending on order
            nik_col = None
            for j in range(i+1, min(i+4, len(headers))):
                hj = (headers[j] or '').strip().upper()
                if 'NIK' in hj or 'NOMOR INDUK' in hj:
                    nik_col = j
                    break
            if nik_col is not None:
                is_cap = player_num == 0
                player_groups.append((i, nik_col, is_cap))
                player_num += 1
                i = nik_col + 1
                continue
        # Also check for "NAMA LENGKAP CADANGAN" pattern
        if 'CADANGAN' in h and 'NAMA' in h:
            for j in range(i+1, min(i+4, len(headers))):
                hj = (headers[j] or '').strip().upper()
                if 'NIK' in hj or 'NOMOR INDUK' in hj:
                    nik_col = j
                    player_groups.append((i, nik_col, False))
                    i = nik_col + 1
                    break
            else:
                i += 1
            continue
        i += 1

    # Edge case: EFO SOLO has single player with column "NAMA LENGKAP PEMAIN"
    if not player_groups:
        # fallback: find any NAMA LENGKAP and NIK
        nama_col = col_idx(headers, 'nama lengkap pemain', 'nama pemain', 'nama lengkap')
        nik_c    = col_idx(headers, 'nomor induk kependudukan (nik)', 'nik')
        if nama_col is not None and nik_c is not None:
            player_groups = [(nama_col, nik_c, True)]

    teams = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        team_name   = clean_str(row[team_col])   if team_col is not None else ''
        school_name = clean_str(row[school_col]) if school_col is not None else ''
        phone       = clean_str(row[phone_col])  if phone_col is not None else ''

        if not team_name or not school_name:
            continue

        players = []
        for (nc, nk, is_cap) in player_groups:
            if nc >= len(row):
                continue
            name = clean_str(row[nc])
            nik  = clean_nik(row[nk]) if nk < len(row) else None
            if name:
                players.append({'name': name, 'nik': nik, 'is_captain': is_cap})

        if players:
            teams.append({
                'team_name': team_name,
                'school_name': school_name,
                'phone': phone,
                'players': players,
            })

    return teams


# ── Main import ────────────────────────────────────────────────────────────────

def main():
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    # Preload game IDs
    cur.execute("SELECT slug, id FROM games")
    game_by_slug = {row[0]: row[1] for row in cur.fetchall()}

    # Caches
    school_cache = {}  # (lower_name, level) → school_id
    team_cache   = {}  # (lower_name, game_id) → team_id
    nik_cache    = {}  # nik → user_id
    email_cache  = set()

    stats = {'schools': 0, 'teams': 0, 'users': 0, 'members': 0,
             'skipped_nik': 0, 'skipped_team': 0, 'errors': []}

    def get_or_create_school(name, level):
        key = (name.lower(), level)
        if key in school_cache:
            return school_cache[key]
        cur.execute("SELECT id FROM schools WHERE lower(name)=%s AND level=%s",
                    (name.lower(), level))
        row = cur.fetchone()
        if row:
            school_cache[key] = row[0]
            return row[0]
        sid = str(uuid.uuid4())
        cur.execute(
            "INSERT INTO schools (id, name, level, city, created_at) VALUES (%s,%s,%s,'Denpasar',now())",
            (sid, name, level)
        )
        stats['schools'] += 1
        school_cache[key] = sid
        return sid

    def get_or_create_team(team_name, game_id, school_id):
        key = (team_name.lower(), game_id)
        if key in team_cache:
            return team_cache[key], False
        cur.execute("SELECT id FROM teams WHERE lower(name)=%s AND game_id=%s",
                    (team_name.lower(), game_id))
        row = cur.fetchone()
        if row:
            team_cache[key] = row[0]
            return row[0], False
        tid = str(uuid.uuid4())
        now = datetime.utcnow()
        cur.execute(
            """INSERT INTO teams (id, name, game_id, school_id, status, created_at, updated_at)
               VALUES (%s,%s,%s,%s,'approved',%s,%s)""",
            (tid, team_name, game_id, school_id, now, now)
        )
        stats['teams'] += 1
        team_cache[key] = tid
        return tid, True

    def get_or_create_user(player, tingkat, nomor_pertandingan):
        nik = player['nik']
        name = player['name']

        # Generate fallback NIK if missing/too short
        if not nik or len(nik) < 8:
            # Use name-based unique key
            slug = re.sub(r'[^a-z0-9]', '', name.lower())[:12]
            base = f"X{slug}"
            counter = 0
            candidate = base
            while True:
                if candidate not in nik_cache:
                    # Also check DB
                    cur.execute("SELECT id FROM users WHERE nisn=%s", (candidate,))
                    if not cur.fetchone():
                        nik = candidate
                        break
                counter += 1
                candidate = f"{base}{counter}"
            stats['skipped_nik'] += 1

        if nik in nik_cache:
            return nik_cache[nik], False

        cur.execute("SELECT id FROM users WHERE nisn=%s", (nik,))
        row = cur.fetchone()
        if row:
            nik_cache[nik] = row[0]
            return row[0], False

        # Generate unique email
        base_email = f"{nik}@porjar.local"
        email = base_email
        counter = 0
        while email in email_cache:
            counter += 1
            email = f"{nik}_{counter}@porjar.local"

        plain = gen_password()
        hashed = hash_password(plain)
        uid = str(uuid.uuid4())
        now = datetime.utcnow()

        cur.execute(
            """INSERT INTO users (id, email, password_hash, full_name, role, nisn, tingkat,
               nomor_pertandingan, needs_password_change, created_at, updated_at)
               VALUES (%s,%s,%s,%s,'player',%s,%s,%s,true,%s,%s)""",
            (uid, email, hashed, name, nik, tingkat, nomor_pertandingan, now, now)
        )
        email_cache.add(email)
        nik_cache[nik] = uid
        stats['users'] += 1
        return uid, True

    def add_team_member(team_id, user_id, name, role):
        cur.execute("SELECT id FROM team_members WHERE team_id=%s AND user_id=%s",
                    (team_id, user_id))
        if cur.fetchone():
            return
        mid = str(uuid.uuid4())
        cur.execute(
            """INSERT INTO team_members (id, team_id, user_id, in_game_name, role, joined_at)
               VALUES (%s,%s,%s,%s,%s,now())""",
            (mid, team_id, user_id, name, role)
        )
        stats['members'] += 1

    def set_captain(team_id, user_id):
        cur.execute("UPDATE teams SET captain_user_id=%s WHERE id=%s AND captain_user_id IS NULL",
                    (user_id, team_id))

    # ── Process each folder ────────────────────────────────────────────────────
    for folder_name, (game_slug, tingkat, nomor_display) in FOLDER_MAP.items():
        folder_path = os.path.join(BASE_DIR, folder_name)
        if not os.path.isdir(folder_path):
            print(f'  [SKIP] Folder tidak ditemukan: {folder_name}')
            continue

        xlsx_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
        if not xlsx_files:
            print(f'  [SKIP] Tidak ada xlsx di: {folder_name}')
            continue

        xlsx_path = os.path.join(folder_path, xlsx_files[0])
        game_id = game_by_slug.get(game_slug)
        if not game_id:
            print(f'  [ERROR] Game slug tidak ditemukan: {game_slug}')
            continue

        print(f'\n→ {folder_name} ({tingkat}, {game_slug})')
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active

        teams_data = extract_players(ws, folder_name)
        print(f'  Parsed {len(teams_data)} tim')

        for td in teams_data:
            try:
                school_id = get_or_create_school(td['school_name'], tingkat)
                team_id, is_new = get_or_create_team(td['team_name'], game_id, school_id)

                captain_set = False
                for player in td['players']:
                    user_id, _ = get_or_create_user(player, tingkat, nomor_display)
                    role = 'captain' if player['is_captain'] else 'member'
                    add_team_member(team_id, user_id, player['name'], role)
                    if player['is_captain'] and not captain_set:
                        set_captain(team_id, user_id)
                        captain_set = True

                conn.commit()
            except Exception as e:
                conn.rollback()
                msg = f"ERROR tim '{td['team_name']}' ({folder_name}): {e}"
                print(f'  !! {msg}')
                stats['errors'].append(msg)

    cur.close()
    conn.close()

    print('\n' + '='*60)
    print(f"Sekolah dibuat  : {stats['schools']}")
    print(f"Tim dibuat      : {stats['teams']}")
    print(f"User dibuat     : {stats['users']}")
    print(f"Member ditambah : {stats['members']}")
    print(f"NIK kosong/skip : {stats['skipped_nik']}")
    if stats['errors']:
        print(f"\nErrors ({len(stats['errors'])}):")
        for e in stats['errors']:
            print(f"  - {e}")

if __name__ == '__main__':
    main()
